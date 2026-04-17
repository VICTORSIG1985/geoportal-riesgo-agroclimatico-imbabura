"""
Script publicado como parte del Geoportal Riesgo Agroclimático — Imbabura, Ecuador
Pinto Páez, V. H. (2026). https://doi.org/10.5281/zenodo.19288559
Licencia: CC BY 4.0

NOTA: este archivo fue sanitizado para publicación pública.
- Credenciales (API keys, tokens, service accounts) removidas → <REDACTED_*>
- Rutas absolutas locales sustituidas por <RUTA_LOCAL>
Antes de ejecutar, configure sus propias rutas de entrada/salida y credenciales.

Archivo original: SCRIPT 04B ADQUISICIÓN MULTI-FUENTE DE DATOS DE DISTRIBUCIÓN DE CULTIVOS.py
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
===============================================================================
SCRIPT 04B: ADQUISICIÓN MULTI-FUENTE DE DATOS DE DISTRIBUCIÓN DE CULTIVOS
===============================================================================
Tesis: Riesgo agroclimático de cultivos andinos bajo escenarios CMIP6
       en la provincia de Imbabura
Autor: Víctor Hugo Pinto Páez
Universidad: Universidad San Gregorio de Portoviejo
Versión: 1.0.0
Fecha: 2026-02-23

JUSTIFICACIÓN:
    El Script 04A documentó que los servicios WFS del MAG Ecuador no
    tienen cobertura geográfica para la Sierra (0/21 capas con datos
    en Imbabura). Este script implementa la estrategia multi-fuente
    para obtener datos de distribución de cultivos:
    
    Fuente 1: ESPAC/INEC 2024 - Estadísticas provinciales (4/4 cultivos)
    Fuente 2: MapSPAM 2020 v2r0 - Distribución espacial global (3/4)
    Fuente 3: CKAN/Datos Abiertos - Shapefiles SIGTIERRAS (intento)
    Fuente 4: PDOTs Prefectura - Datos parroquiales (descarga PDF)

REFERENCIAS:
    - INEC Ecuador. ESPAC 2024. ecuadorencifras.gob.ec
    - IFPRI. MapSPAM 2020 v2r0. doi:10.7910/DVN/SWPENT
    - MAG Ecuador. SIGTIERRAS. sigtierras.gob.ec
    - Prefectura de Imbabura. PDOTs 2023-2027. imbabura.gob.ec
===============================================================================
"""

import os
import sys
import json
import logging
import zipfile
import shutil
import time
from datetime import datetime
from pathlib import Path

# ── Dependencias ──────────────────────────────────────────────────────
DEPS_FALTANTES = []
try:
    import requests
except ImportError:
    DEPS_FALTANTES.append("requests")
try:
    import numpy as np
except ImportError:
    DEPS_FALTANTES.append("numpy")
try:
    import pandas as pd
except ImportError:
    DEPS_FALTANTES.append("pandas")
try:
    import geopandas as gpd
except ImportError:
    DEPS_FALTANTES.append("geopandas")
try:
    import rasterio
    from rasterio.mask import mask as rasterio_mask
except ImportError:
    DEPS_FALTANTES.append("rasterio")
try:
    from shapely.geometry import box
except ImportError:
    DEPS_FALTANTES.append("shapely")
try:
    import openpyxl
except ImportError:
    DEPS_FALTANTES.append("openpyxl")

if DEPS_FALTANTES:
    print(f"ERROR: Dependencias faltantes: {', '.join(DEPS_FALTANTES)}")
    print(f"Instalar con: pip install {' '.join(DEPS_FALTANTES)}")
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════════════

TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
FECHA_ISO = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Rutas
BASE_DIR = Path(r"<RUTA_LOCAL>")
DATOS_DIR = BASE_DIR / "02_DATOS"
CULTIVOS_RAW = DATOS_DIR / "cultivos" / "raw"
CULTIVOS_PROC = DATOS_DIR / "cultivos" / "procesados"
ESPAC_DIR = CULTIVOS_RAW / "ESPAC"
MAPSPAM_DIR = CULTIVOS_RAW / "MapSPAM"
SIGTIERRAS_DIR = CULTIVOS_RAW / "SIGTIERRAS"
PDOT_DIR = CULTIVOS_RAW / "PDOTs"
REPORTES_DIR = BASE_DIR / "05_DOCUMENTACION" / "reportes_auditoria"
DOC_DIR = BASE_DIR / "05_DOCUMENTACION"

# ── Cultivos objetivo con mapeo directo a tablas ESPAC ────────────────
# El ÍNDICE del archivo Tabulados_ESPAC_2024.xlsx (INEC) identifica
# cada tabla por cultivo específico y forma de comercialización.
CULTIVOS = {
    'papa': {
        'nombre_cientifico': 'Solanum tuberosum L.',
        'nombre_comun': 'Papa',
        'espac_tablas': {'T43': 'Papa (tubérculo fresco)'},
        'mapspam_code': 'POTA',
        'mapspam_disponible': True,
    },
    'maiz': {
        'nombre_cientifico': 'Zea mays L.',
        'nombre_comun': 'Maíz',
        'espac_tablas': {
            'T38': 'Maíz duro choclo (en choclo)',
            'T39': 'Maíz duro seco (grano seco)',
            'T40': 'Maíz suave choclo (en choclo)',
            'T41': 'Maíz suave seco (grano seco)',
        },
        'mapspam_code': 'MAIZ',
        'mapspam_disponible': True,
    },
    'frejol': {
        'nombre_cientifico': 'Phaseolus vulgaris L.',
        'nombre_comun': 'Fréjol',
        'espac_tablas': {
            'T34': 'Fréjol seco (grano seco)',
            'T35': 'Fréjol tierno (en vaina)',
        },
        'mapspam_code': 'BEAN',
        'mapspam_disponible': True,
    },
    'quinua': {
        'nombre_cientifico': 'Chenopodium quinoa Willd.',
        'nombre_comun': 'Quinua',
        'espac_tablas': {'T44': 'Quinua (grano seco)'},
        'mapspam_code': None,
        'mapspam_disponible': False,
    },
}

# URLs
URLS = {
    'espac_tabulados': 'https://www.ecuadorencifras.gob.ec/documentos/web-inec/Estadisticas_agropecuarias/espac/2024/Tabulados_ESPAC_2024.xlsx',
    'espac_series': 'https://www.ecuadorencifras.gob.ec/documentos/web-inec/Estadisticas_agropecuarias/espac/2024/Series_historicas_2014-2024.xlsx',
    'mapspam_harvested': 'https://www.dropbox.com/scl/fi/vgxfy41otygcee89apst0/spam2020V2r0_global_harvested_area.geotiff.zip?rlkey=esz9aoh6f79zorhmv9zwrlnpx&dl=1',
    'ckan_cobertura': 'https://datosabiertos.gob.ec/api/3/action/package_show?id=mapa-de-cobertura-y-uso-de-la-tierra',
    'pdot_provincial': 'https://www.imbabura.gob.ec/phocadownloadpap/K-Planes-programas/PDOT/pdot_resumido_provincia_de_imbabura_2023-2027.pdf',
}

# BBOX Imbabura WGS84 (con buffer 10km, del Script 02)
IMBABURA_BBOX = {
    'lon_min': -79.3649, 'lon_max': -77.7205,
    'lat_min': 0.0307,   'lat_max': 0.9669,
}


# ═══════════════════════════════════════════════════════════════════════
# LOGGING
# ═══════════════════════════════════════════════════════════════════════

def configurar_logging():
    """Configurar logging dual (consola + archivo)."""
    log_file = REPORTES_DIR / f"SCRIPT_04B_LOG_{TIMESTAMP}.txt"
    logger = logging.getLogger('SCRIPT_04B')
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fmt = logging.Formatter('%(asctime)s | %(levelname)-8s | %(message)s',
                            datefmt='%Y-%m-%d %H:%M:%S')
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    logger.addHandler(ch)
    fh = logging.FileHandler(log_file, encoding='utf-8')
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    return logger, log_file


# ═══════════════════════════════════════════════════════════════════════
# FUNCIONES AUXILIARES
# ═══════════════════════════════════════════════════════════════════════

def crear_directorios():
    """Crear estructura de directorios."""
    for d in [CULTIVOS_RAW, CULTIVOS_PROC, ESPAC_DIR, MAPSPAM_DIR,
              SIGTIERRAS_DIR, PDOT_DIR, REPORTES_DIR, DOC_DIR]:
        d.mkdir(parents=True, exist_ok=True)


def descargar_archivo(url, destino, descripcion, logger, timeout=300):
    """Descargar archivo con reintentos (3 intentos)."""
    resultado = {'exito': False, 'size_mb': 0.0, 'tiempo_s': 0.0, 'error': ''}
    for intento in range(3):
        try:
            t0 = time.time()
            logger.info(f"  Descargando: {descripcion} (intento {intento+1}/3)")
            r = requests.get(url, stream=True, timeout=timeout,
                             allow_redirects=True,
                             headers={'User-Agent': 'Mozilla/5.0 (TesisImb/1.0)'})
            r.raise_for_status()
            total = int(r.headers.get('content-length', 0))
            descargado = 0
            with open(destino, 'wb') as f:
                for chunk in r.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
                        descargado += len(chunk)
                        if total > 0 and descargado % (8192 * 100) == 0:
                            logger.info(f"    Progreso: {descargado/total*100:.0f}%"
                                       f" ({descargado/1e6:.1f}/{total/1e6:.1f} MB)")
            resultado['tiempo_s'] = time.time() - t0
            resultado['size_mb'] = destino.stat().st_size / 1e6
            resultado['exito'] = True
            logger.info(f"  ✓ {resultado['size_mb']:.2f} MB en {resultado['tiempo_s']:.1f}s")
            return resultado
        except requests.exceptions.Timeout:
            resultado['error'] = f"Timeout ({timeout}s)"
            logger.warning(f"  ⚠ Timeout intento {intento+1}")
        except requests.exceptions.HTTPError as e:
            resultado['error'] = f"HTTP {e.response.status_code}"
            logger.warning(f"  ⚠ {resultado['error']}")
            break
        except Exception as e:
            resultado['error'] = str(e)[:200]
            logger.warning(f"  ⚠ {resultado['error']}")
        if intento < 2:
            time.sleep(5 * (intento + 1))
    logger.error(f"  ✗ Falló: {descripcion} → {resultado['error']}")
    return resultado


# ═══════════════════════════════════════════════════════════════════════
# FUENTE 1: ESPAC/INEC
# ═══════════════════════════════════════════════════════════════════════

def descargar_y_procesar_espac(logger):
    """
    Descargar y procesar datos ESPAC 2024 del INEC Ecuador.
    
    ESPAC (Encuesta de Superficie y Producción Agropecuaria Continua)
    proporciona estadísticas de superficie sembrada, cosechada, producción
    y rendimiento a nivel provincial. Es la única fuente que cubre los
    4 cultivos objetivo incluyendo quinua.
    
    El procesamiento usa mapeo directo del ÍNDICE del Excel para
    identificar las tablas específicas por cultivo:
      T34/T35 = Fréjol, T38-T41 = Maíz, T43 = Papa, T44 = Quinua
    
    Estructura de cada tabla ESPAC:
      Fila 10: Headers (Región y Provincia | | SUPERFICIE | | PRODUCCIÓN | VENTAS)
      Fila 11: Sub-headers (| | Plantada | Cosechada | | )
      Fila 12: Total Nacional
      Fila 13+: Datos por región/provincia
    
    Referencia: INEC (2024). ESPAC 2024.
    https://www.ecuadorencifras.gob.ec/estadisticas-agropecuarias-2/
    """
    logger.info("=" * 60)
    logger.info("FUENTE 1: ESPAC/INEC - Estadísticas provinciales")
    logger.info("=" * 60)
    
    resultados = {}
    
    # ── Descargar ─────────────────────────────────────────────────────
    dest_tab = ESPAC_DIR / "Tabulados_ESPAC_2024.xlsx"
    if not dest_tab.exists():
        resultados['tabulados'] = descargar_archivo(
            URLS['espac_tabulados'], dest_tab, "ESPAC 2024 - Tabulados", logger, 60)
    else:
        logger.info(f"  → Ya existe: {dest_tab.name}")
        resultados['tabulados'] = {'exito': True, 'size_mb': dest_tab.stat().st_size/1e6,
                                    'tiempo_s': 0, 'error': ''}
    
    dest_series = ESPAC_DIR / "Series_historicas_2014-2024.xlsx"
    if not dest_series.exists():
        resultados['series'] = descargar_archivo(
            URLS['espac_series'], dest_series, "ESPAC Series 2014-2024", logger, 60)
    else:
        logger.info(f"  → Ya existe: {dest_series.name}")
        resultados['series'] = {'exito': True, 'size_mb': dest_series.stat().st_size/1e6,
                                 'tiempo_s': 0, 'error': ''}
    
    if not resultados.get('tabulados', {}).get('exito'):
        return resultados
    
    # ── Procesar: Mapeo directo por ÍNDICE del Excel ──────────────────
    logger.info("")
    logger.info("  Procesando datos ESPAC para Imbabura...")
    
    try:
        wb = openpyxl.load_workbook(dest_tab, data_only=True)
        datos_cultivos = {}
        
        for cultivo, config in CULTIVOS.items():
            logger.info(f"\n  ── {config['nombre_comun']} ({config['nombre_cientifico']}) ──")
            subtablas = []
            
            for tabla_id, tabla_desc in config['espac_tablas'].items():
                if tabla_id not in wb.sheetnames:
                    logger.warning(f"    ⚠ Hoja '{tabla_id}' no encontrada")
                    continue
                
                ws = wb[tabla_id]
                logger.info(f"    Tabla {tabla_id}: {tabla_desc}")
                
                # Buscar fila de Imbabura
                fila_imbabura = None
                fila_num = None
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    for cell in row:
                        if cell and 'imbabura' in str(cell).lower():
                            fila_imbabura = row
                            fila_num = row_idx
                            break
                    if fila_imbabura:
                        break
                
                if fila_imbabura is None:
                    logger.warning(f"      ⚠ Imbabura no encontrada en {tabla_id}")
                    continue
                
                logger.info(f"      Encontrado en fila {fila_num}")
                
                # Extraer valores numéricos después de "Imbabura"
                numericos = []
                pasado_imbabura = False
                for v in fila_imbabura:
                    if v and 'imbabura' in str(v).lower():
                        pasado_imbabura = True
                        continue
                    if pasado_imbabura and isinstance(v, (int, float)) and not isinstance(v, bool):
                        numericos.append(float(v))
                    elif pasado_imbabura and isinstance(v, str) and v.strip().lower() in ('solo', 'asociado'):
                        continue
                
                dato = {
                    'tabla': tabla_id,
                    'descripcion': tabla_desc,
                    'sup_plantada_ha': numericos[0] if len(numericos) > 0 else None,
                    'sup_cosechada_ha': numericos[1] if len(numericos) > 1 else None,
                    'produccion_tm': numericos[2] if len(numericos) > 2 else None,
                    'ventas_tm': numericos[3] if len(numericos) > 3 else None,
                }
                
                for campo, label, unidad in [
                    ('sup_plantada_ha', 'Sup. plantada', 'ha'),
                    ('sup_cosechada_ha', 'Sup. cosechada', 'ha'),
                    ('produccion_tm', 'Producción', 'Tm'),
                    ('ventas_tm', 'Ventas', 'Tm')]:
                    val = dato[campo]
                    logger.info(f"      {label:18s}: {val:>12.2f} {unidad}" if val else f"      {label:18s}: N/D")
                
                subtablas.append(dato)
            
            if subtablas:
                total = {
                    'nombre_comun': config['nombre_comun'],
                    'nombre_cientifico': config['nombre_cientifico'],
                    'subtablas': subtablas,
                    'total_sup_plantada_ha': sum(d['sup_plantada_ha'] or 0 for d in subtablas),
                    'total_sup_cosechada_ha': sum(d['sup_cosechada_ha'] or 0 for d in subtablas),
                    'total_produccion_tm': sum(d['produccion_tm'] or 0 for d in subtablas),
                    'total_ventas_tm': sum(d['ventas_tm'] or 0 for d in subtablas),
                }
                datos_cultivos[cultivo] = total
                logger.info(f"    ► TOTAL {config['nombre_comun'].upper()}: "
                           f"{total['total_sup_plantada_ha']:.1f} ha, "
                           f"{total['total_produccion_tm']:.1f} Tm")
        
        # ── JSON ──────────────────────────────────────────────────────
        salida = {
            'fuente': 'ESPAC/INEC 2024 - Tabulados',
            'fecha_procesamiento': FECHA_ISO,
            'metodo': 'Mapeo directo por ÍNDICE del Excel a tablas por cultivo',
            'provincia': 'Imbabura',
            'resolucion': 'Provincial',
            'cultivos': {},
        }
        for cult, datos in datos_cultivos.items():
            salida['cultivos'][cult] = {
                'nombre_comun': datos['nombre_comun'],
                'nombre_cientifico': datos['nombre_cientifico'],
                'sup_plantada_ha': round(datos['total_sup_plantada_ha'], 2),
                'sup_cosechada_ha': round(datos['total_sup_cosechada_ha'], 2),
                'produccion_tm': round(datos['total_produccion_tm'], 2),
                'ventas_tm': round(datos['total_ventas_tm'], 2),
                'detalle_por_subtipo': [
                    {k: (round(v, 2) if isinstance(v, float) else v) for k, v in d.items()}
                    for d in datos['subtablas']
                ]
            }
        
        json_path = CULTIVOS_PROC / "espac_imbabura_resumen.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(salida, f, indent=2, ensure_ascii=False)
        logger.info(f"\n  ✓ JSON: {json_path.name}")
        
        # ── CSV ───────────────────────────────────────────────────────
        csv_data = []
        for cult, datos in datos_cultivos.items():
            csv_data.append({
                'cultivo': cult,
                'nombre_comun': datos['nombre_comun'],
                'sup_plantada_ha': round(datos['total_sup_plantada_ha'], 2),
                'sup_cosechada_ha': round(datos['total_sup_cosechada_ha'], 2),
                'produccion_tm': round(datos['total_produccion_tm'], 2),
                'ventas_tm': round(datos['total_ventas_tm'], 2),
            })
        df = pd.DataFrame(csv_data)
        csv_path = CULTIVOS_PROC / "espac_imbabura_cultivos.csv"
        df.to_csv(csv_path, index=False, encoding='utf-8')
        logger.info(f"  ✓ CSV: {csv_path.name}")
        
        # ── Tabla resumen ─────────────────────────────────────────────
        logger.info(f"\n  {'='*65}")
        logger.info(f"  RESUMEN ESPAC 2024 - IMBABURA")
        logger.info(f"  {'─'*65}")
        logger.info(f"  {'Cultivo':<12} {'Sup.Plant.(ha)':>15} {'Sup.Cos.(ha)':>14} {'Prod.(Tm)':>12}")
        logger.info(f"  {'─'*65}")
        for _, row in df.iterrows():
            logger.info(f"  {row['nombre_comun']:<12} {row['sup_plantada_ha']:>15.1f} "
                       f"{row['sup_cosechada_ha']:>14.1f} {row['produccion_tm']:>12.1f}")
        logger.info(f"  {'─'*65}")
        logger.info(f"  {'TOTAL':<12} {df['sup_plantada_ha'].sum():>15.1f} "
                   f"{df['sup_cosechada_ha'].sum():>14.1f} {df['produccion_tm'].sum():>12.1f}")
        logger.info(f"  {'='*65}")
        
        resultados['procesamiento'] = {
            'exito': True,
            'cultivos_encontrados': list(datos_cultivos.keys()),
            'n_cultivos': len(datos_cultivos),
        }
        
    except Exception as e:
        logger.error(f"  ✗ Error procesando ESPAC: {e}")
        resultados['procesamiento'] = {'exito': False, 'error': str(e)}
    
    return resultados


# ═══════════════════════════════════════════════════════════════════════
# FUENTE 2: MapSPAM 2020
# ═══════════════════════════════════════════════════════════════════════

def descargar_y_procesar_mapspam(logger):
    """
    Descargar y procesar MapSPAM 2020 v2r0 (IFPRI).
    
    Datos espacializados de superficie cosechada a ~10 km. Se usa la
    variante _A (All technologies) que integra todos los sistemas
    productivos. Disponible para papa (POTA), maíz (MAIZ), fréjol (BEAN).
    Quinua no está mapeada individualmente en MapSPAM.
    
    Nomenclatura: spam2020V2r0_global_H_{CROP}_A.tif
      H = Harvested area, A = All technologies
    
    Referencia: IFPRI (2024). MapSPAM 2020 v2r0.
    https://doi.org/10.7910/DVN/SWPENT
    """
    logger.info("")
    logger.info("=" * 60)
    logger.info("FUENTE 2: MapSPAM 2020 - Distribución espacial global")
    logger.info("=" * 60)
    logger.info("  Variante: _A (All technologies)")
    logger.info("  Quinua: NO disponible en MapSPAM")
    
    resultados = {}
    
    # ── Descargar ─────────────────────────────────────────────────────
    dest_zip = MAPSPAM_DIR / "spam2020V2r0_global_harvested_area.geotiff.zip"
    if not dest_zip.exists():
        resultados['descarga'] = descargar_archivo(
            URLS['mapspam_harvested'], dest_zip,
            "MapSPAM 2020 Harvested Area (~65 MB)", logger, 600)
    else:
        logger.info(f"  → Ya existe: {dest_zip.name} ({dest_zip.stat().st_size/1e6:.1f} MB)")
        resultados['descarga'] = {'exito': True, 'size_mb': dest_zip.stat().st_size/1e6,
                                   'tiempo_s': 0, 'error': ''}
    
    if not resultados.get('descarga', {}).get('exito'):
        return resultados
    
    # ── Extraer ───────────────────────────────────────────────────────
    extract_dir = MAPSPAM_DIR / "extracted"
    if not extract_dir.exists():
        logger.info(f"\n  Extrayendo ZIP...")
        with zipfile.ZipFile(dest_zip, 'r') as zf:
            zf.extractall(extract_dir)
            logger.info(f"  ✓ {len(zf.namelist())} archivos extraídos")
    
    # ── Recortar a Imbabura ───────────────────────────────────────────
    logger.info("  Recortando a BBOX de Imbabura...")
    bbox_geom = box(IMBABURA_BBOX['lon_min'], IMBABURA_BBOX['lat_min'],
                    IMBABURA_BBOX['lon_max'], IMBABURA_BBOX['lat_max'])
    
    cultivos_mapspam = {c: v['mapspam_code'] for c, v in CULTIVOS.items()
                        if v['mapspam_disponible']}
    
    proc = {'exito': False, 'cultivos_procesados': [], 'archivos': []}
    
    for cultivo, code in cultivos_mapspam.items():
        logger.info(f"\n  ── {cultivo.upper()} (código: {code}) ──")
        
        # Buscar SOLO la variante _A
        matches = [f for f in extract_dir.rglob("*.tif")
                   if code in f.name.upper() and '_A.' in f.name]
        if not matches:
            matches = [f for f in extract_dir.rglob("*.tif")
                       if f"H_{code}_A" in f.name.upper()]
        if not matches:
            logger.warning(f"    ⚠ No encontrado: {code}_A.tif")
            continue
        
        tif_path = matches[0]
        logger.info(f"    Fuente: {tif_path.name}")
        out_path = CULTIVOS_PROC / f"mapspam2020_harvested_{cultivo}_imbabura.tif"
        
        try:
            with rasterio.open(tif_path) as src:
                out_image, out_transform = rasterio_mask(
                    src, [bbox_geom], crop=True, nodata=src.nodata or -9999)
                meta = src.meta.copy()
                meta.update({'height': out_image.shape[1], 'width': out_image.shape[2],
                             'transform': out_transform, 'compress': 'lzw'})
                with rasterio.open(out_path, 'w', **meta) as dst:
                    dst.write(out_image)
                
                data = out_image[0]
                nodata_val = src.nodata if src.nodata else -9999
                valid = data[(data != nodata_val) & (data > 0) & (~np.isnan(data))]
                
                logger.info(f"    ✓ {out_path.name}: {out_image.shape[2]}x{out_image.shape[1]} px")
                if len(valid) > 0:
                    logger.info(f"      Área: min={valid.min():.1f}, max={valid.max():.1f}, "
                               f"mean={valid.mean():.1f}, total={valid.sum():.1f} ha")
                    logger.info(f"      Píxeles con datos: {len(valid)}")
                else:
                    logger.warning(f"      ⚠ Sin datos válidos en BBOX")
                
                proc['cultivos_procesados'].append(cultivo)
                proc['archivos'].append(str(out_path))
        except Exception as e:
            logger.error(f"    ✗ Error: {e}")
    
    proc['exito'] = len(proc['cultivos_procesados']) > 0
    resultados['procesamiento'] = proc
    
    # Resumen JSON
    mj = CULTIVOS_PROC / "mapspam_imbabura_resumen.json"
    with open(mj, 'w', encoding='utf-8') as f:
        json.dump({
            'fuente': 'MapSPAM 2020 v2r0 (IFPRI)', 'doi': '10.7910/DVN/SWPENT',
            'variante': 'H_*_A (Harvested Area, All technologies)',
            'fecha_procesamiento': FECHA_ISO,
            'cultivos_procesados': proc['cultivos_procesados'],
            'archivos': proc['archivos'],
            'nota_quinua': 'No mapeada individualmente en MapSPAM',
        }, f, indent=2, ensure_ascii=False)
    logger.info(f"\n  ✓ Resumen: {mj.name}")
    
    return resultados


# ═══════════════════════════════════════════════════════════════════════
# FUENTE 3: CKAN / DATOS ABIERTOS ECUADOR
# ═══════════════════════════════════════════════════════════════════════

def intentar_ckan(logger):
    """
    Intentar descarga de shapefiles SIGTIERRAS via API CKAN.
    Fallback: instrucciones de descarga manual.
    
    Referencia: Datos Abiertos Ecuador. datosabiertos.gob.ec
    """
    logger.info("")
    logger.info("=" * 60)
    logger.info("FUENTE 3: CKAN/Datos Abiertos - Shapefiles SIGTIERRAS")
    logger.info("=" * 60)
    
    resultado = {'exito': False, 'error': '', 'recursos': []}
    
    try:
        logger.info("  Consultando API CKAN (timeout 30s)...")
        r = requests.get(URLS['ckan_cobertura'], timeout=30,
                         headers={'User-Agent': 'Mozilla/5.0 (TesisImb/1.0)'})
        r.raise_for_status()
        data = r.json()
        if data.get('success'):
            recursos = data['result']['resources']
            logger.info(f"  ✓ {len(recursos)} recursos encontrados")
            for rec in recursos:
                logger.info(f"    - {rec['name'][:60]}: {rec['format']}")
                resultado['recursos'].append({'nombre': rec['name'], 'formato': rec['format'],
                                               'url': rec['url']})
            shps = [rec for rec in recursos if rec['format'].upper() in ('SHP', 'ZIP', 'SHAPEFILE')]
            if shps:
                for rec in shps[:2]:
                    dest = SIGTIERRAS_DIR / f"ckan_{rec['name'][:50]}.{rec['format'].lower()}"
                    dl = descargar_archivo(rec['url'], dest, f"CKAN: {rec['name'][:40]}", logger, 120)
                    if dl['exito']:
                        resultado['exito'] = True
                        break
    except requests.exceptions.Timeout:
        resultado['error'] = "Timeout CKAN (30s) - servicio no disponible"
        logger.warning(f"  ⚠ {resultado['error']}")
    except Exception as e:
        resultado['error'] = str(e)[:200]
        logger.warning(f"  ⚠ {resultado['error']}")
    
    if not resultado['exito']:
        logger.info("")
        logger.info("  DESCARGA MANUAL REQUERIDA:")
        logger.info("  1. http://www.sigtierras.gob.ec/descargas/")
        logger.info("     → Cartografía Temática → Cobertura y Uso → Imbabura")
        logger.info("  2. https://datosabiertos.gob.ec/dataset/mapa-de-cobertura-y-uso-de-la-tierra")
        logger.info(f"  Destino: {SIGTIERRAS_DIR}")
    
    return resultado


# ═══════════════════════════════════════════════════════════════════════
# FUENTE 4: PDOTs PREFECTURA DE IMBABURA
# ═══════════════════════════════════════════════════════════════════════

def descargar_pdots(logger):
    """
    Descargar PDOTs de la Prefectura de Imbabura.
    
    Los PDOTs contienen tablas de superficie cultivada por tipo de cultivo
    a nivel parroquial. Es la única fuente con datos parroquiales para
    quinua. Formato PDF: extracción de tablas en Script 04C.
    
    Referencia: Prefectura de Imbabura (2023). PDOT 2023-2027.
    """
    logger.info("")
    logger.info("=" * 60)
    logger.info("FUENTE 4: PDOTs - Prefectura de Imbabura")
    logger.info("=" * 60)
    
    resultados = {}
    pdots = {
        'provincial': (
            URLS['pdot_provincial'],
            PDOT_DIR / "PDOT_Provincial_Imbabura_2023-2027.pdf",
            "PDOT Provincial 2023-2027"),
        'cantonal_Ibarra': (
            'https://www.imbabura.gob.ec/phocadownloadpap/K-Planes-programas/PDOT/Cantonal/PDOT%20IBARRA.pdf',
            PDOT_DIR / "PDOT_Ibarra.pdf",
            "PDOT Ibarra"),
        'cantonal_Cotacachi': (
            'https://www.imbabura.gob.ec/phocadownloadpap/K-Planes-programas/PDOT/Cantonal/PDOT%20COTACACHI.pdf',
            PDOT_DIR / "PDOT_Cotacachi.pdf",
            "PDOT Cotacachi"),
        'cantonal_Urcuqui': (
            'https://www.imbabura.gob.ec/phocadownloadpap/K-Planes-programas/PDOT/Cantonal/PDOT%20SAN%20MIGUEL%20DE%20URCUQUI%CC%81.pdf',
            PDOT_DIR / "PDOT_Urcuqui.pdf",
            "PDOT Urcuquí"),
    }
    
    for key, (url, dest, desc) in pdots.items():
        if not dest.exists():
            resultados[key] = descargar_archivo(url, dest, desc, logger, 120)
        else:
            logger.info(f"  → Ya existe: {dest.name} ({dest.stat().st_size/1e6:.1f} MB)")
            resultados[key] = {'exito': True, 'size_mb': dest.stat().st_size/1e6,
                                'tiempo_s': 0, 'error': ''}
    
    logger.info("")
    logger.info("  NOTA: Datos en PDF. Extracción de tablas en Script 04C.")
    return resultados


# ═══════════════════════════════════════════════════════════════════════
# REPORTE
# ═══════════════════════════════════════════════════════════════════════

def generar_reporte(resultados, t_inicio, logger):
    """Generar reporte de auditoría."""
    t_total = time.time() - t_inicio
    reporte_path = REPORTES_DIR / f"REPORTE_SCRIPT_04B_{TIMESTAMP}.txt"
    
    espac = resultados.get('espac', {})
    espac_proc = espac.get('procesamiento', {})
    mapspam = resultados.get('mapspam', {})
    mapspam_proc = mapspam.get('procesamiento', {})
    ckan = resultados.get('ckan', {})
    pdots = resultados.get('pdots', {})
    pdots_ok = sum(1 for v in pdots.values() if isinstance(v, dict) and v.get('exito'))
    
    lines = [
        "=" * 70,
        "REPORTE DE AUDITORÍA - SCRIPT 04B: ADQUISICIÓN MULTI-FUENTE",
        "=" * 70,
        "",
        "INFORMACIÓN GENERAL",
        "-" * 30,
        f"Fecha de ejecución: {FECHA_ISO}",
        f"Versión del script: 1.0.0",
        f"Autor: Víctor Hugo Pinto Páez",
        "",
        "RESULTADOS POR FUENTE",
        "-" * 30,
        "",
        "  FUENTE 1: ESPAC/INEC 2024",
        f"    Tabulados: {'✓' if espac.get('tabulados', {}).get('exito') else '✗'}"
        f" ({espac.get('tabulados', {}).get('size_mb', 0):.2f} MB)",
        f"    Series históricas: {'✓' if espac.get('series', {}).get('exito') else '✗'}"
        f" ({espac.get('series', {}).get('size_mb', 0):.2f} MB)",
        f"    Cultivos: {', '.join(espac_proc.get('cultivos_encontrados', []))}",
        "",
        "  FUENTE 2: MapSPAM 2020 v2r0",
        f"    Descarga: {'✓' if mapspam.get('descarga', {}).get('exito') else '✗'}"
        f" ({mapspam.get('descarga', {}).get('size_mb', 0):.1f} MB)",
        f"    Variante: _A (All technologies)",
        f"    Cultivos recortados: {', '.join(mapspam_proc.get('cultivos_procesados', []))}",
        f"    Quinua: NO DISPONIBLE (no mapeada individualmente)",
        "",
        "  FUENTE 3: CKAN/Datos Abiertos",
        f"    Estado: {'✓' if ckan.get('exito') else '✗ No disponible'}",
    ]
    if ckan.get('error'):
        lines.append(f"    Error: {ckan['error']}")
    lines.extend([
        "",
        "  FUENTE 4: PDOTs Prefectura de Imbabura",
        f"    Descargados: {pdots_ok}/{len(pdots)}",
        f"    Extracción de tablas: Pendiente (Script 04C)",
        "",
        "RESUMEN CONSOLIDADO POR CULTIVO",
        "-" * 30,
    ])
    
    for cultivo, config in CULTIVOS.items():
        fuentes = []
        if cultivo in espac_proc.get('cultivos_encontrados', []):
            tablas = ', '.join(config['espac_tablas'].keys())
            fuentes.append(f"ESPAC ({tablas})")
        if cultivo in mapspam_proc.get('cultivos_procesados', []):
            fuentes.append("MapSPAM")
        if pdots_ok > 0:
            fuentes.append("PDOTs")
        lines.append(f"  {'✓' if fuentes else '✗'} {config['nombre_comun']:8s}: "
                    f"{', '.join(fuentes) if fuentes else 'Sin datos'}")
    
    lines.extend([
        "",
        f"TIEMPO DE EJECUCIÓN: {t_total:.1f}s ({t_total/60:.1f} min)",
        "",
        "VERIFICACIÓN DE CRITERIOS",
        "-" * 30,
        f"[{'✓' if espac_proc.get('n_cultivos', 0) == 4 else '✗'}] ESPAC: 4/4 cultivos con datos para Imbabura",
        f"[{'✓' if mapspam_proc.get('exito') else '✗'}] MapSPAM: 3 cultivos recortados a Imbabura",
        f"[{'✓' if ckan.get('exito') else '⚠'}] CKAN/SIGTIERRAS: "
        f"{'disponible' if ckan.get('exito') else 'requiere descarga manual'}",
        f"[{'✓' if pdots_ok > 0 else '✗'}] PDOTs: {pdots_ok} documentos descargados",
        f"[✓] Brecha quinua documentada (solo ESPAC + PDOTs)",
        "",
        "DECISIÓN",
        "-" * 30,
    ])
    
    n_fuentes = sum([1 if espac_proc.get('exito') else 0,
                     1 if mapspam_proc.get('exito') else 0,
                     1 if ckan.get('exito') else 0,
                     1 if pdots_ok > 0 else 0])
    
    if n_fuentes >= 2:
        lines.append("APROBADO - Datos suficientes para proceder.")
        lines.append("Siguiente paso: Script 04C (Integración y agregación parroquial)")
    else:
        lines.append("PENDIENTE - Se requieren fuentes adicionales.")
    
    lines.extend(["", "=" * 70])
    texto = "\n".join(lines)
    
    with open(reporte_path, 'w', encoding='utf-8') as f:
        f.write(texto)
    logger.info("")
    logger.info(texto)
    logger.info(f"\nReporte guardado: {reporte_path}")
    return reporte_path


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    t_inicio = time.time()
    crear_directorios()
    logger, log_file = configurar_logging()
    
    logger.info("=" * 70)
    logger.info("SCRIPT 04B: ADQUISICIÓN MULTI-FUENTE DE DATOS DE CULTIVOS")
    logger.info("=" * 70)
    logger.info(f"Fecha: {FECHA_ISO}")
    logger.info(f"Justificación: Script 04A documentó 0 capas WFS con datos en Imbabura")
    logger.info("")
    
    resultados = {}
    
    for nombre, funcion in [('espac', descargar_y_procesar_espac),
                            ('mapspam', descargar_y_procesar_mapspam),
                            ('ckan', intentar_ckan),
                            ('pdots', descargar_pdots)]:
        try:
            resultados[nombre] = funcion(logger)
        except Exception as e:
            logger.error(f"ERROR CRÍTICO en {nombre}: {e}")
            resultados[nombre] = {'error': str(e)}
    
    generar_reporte(resultados, t_inicio, logger)
    
    # Estado
    estado_path = CULTIVOS_PROC / f"estado_04B_{TIMESTAMP}.json"
    estado = {'version': '1.0.0', 'fecha': FECHA_ISO}
    for fuente, datos in resultados.items():
        if isinstance(datos, dict):
            estado[fuente] = {k: v for k, v in datos.items()
                              if isinstance(v, (str, int, float, bool, list, dict))}
    with open(estado_path, 'w', encoding='utf-8') as f:
        json.dump(estado, f, indent=2, ensure_ascii=False, default=str)
    logger.info(f"Estado guardado: {estado_path}")
    
    t_total = time.time() - t_inicio
    logger.info("")
    logger.info("=" * 70)
    logger.info(f"✓ SCRIPT 04B COMPLETADO en {t_total:.1f}s ({t_total/60:.1f} min)")
    logger.info(f"  Siguiente paso: Script 04C - Integración parroquial")
    logger.info("=" * 70)


if __name__ == '__main__':
    main()