"""
Script publicado como parte del Geoportal Riesgo Agroclimático — Imbabura, Ecuador
Pinto Páez, V. H. (2026). https://doi.org/10.5281/zenodo.19288559
Licencia: CC BY 4.0

NOTA: este archivo fue sanitizado para publicación pública.
- Credenciales (API keys, tokens, service accounts) removidas → <REDACTED_*>
- Rutas absolutas locales sustituidas por <RUTA_LOCAL>
Antes de ejecutar, configure sus propias rutas de entrada/salida y credenciales.

Archivo original: SCRIPT 04C DESAGREGACIÓN ESPACIAL DE SUPERFICIE AGRÍCOLA POR PARROQUIA.py
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 SCRIPT 04C: DESAGREGACIÓN ESPACIAL DE SUPERFICIE AGRÍCOLA POR PARROQUIA
================================================================================

 Tesis:        Riesgo agroclimático de cultivos andinos bajo escenarios CMIP6
               en la provincia de Imbabura: modelamiento de distribución de
               especies para la gestión territorial

 Autor:        Víctor Hugo Pinto Páez
 Universidad:  Universidad San Gregorio de Portoviejo
 Programa:     Maestría en Prevención y Gestión de Riesgos - Mención en
               Variabilidad Climática y Resiliencia Territorial

 Versión:      3.0.0
 Fecha:        2026-02-24

================================================================================
 CAMBIOS RESPECTO A v3.0.0
================================================================================

 CORRECCIÓN CRÍTICA v3.0: Reemplazado rasterstats por exactextract para
    estadísticas zonales.

    PROBLEMA IDENTIFICADO: rasterstats ofrece dos modos para asignar
    píxeles a polígonos, ambos incorrectos cuando la resolución del raster
    (~86 km²/píxel) es mayor que las unidades de análisis (parroquias de
    10-886 km²):

    - all_touched=True: Un píxel compartido entre N parroquias suma su
      valor completo en las N. Diagnóstico demostró inflación de +215%
      a +387% por doble conteo (42,395 ha vs 13,469 ha reales para maíz).

    - all_touched=False: Solo cuenta píxeles cuyo centroide cae dentro
      del polígono. Parroquias menores que un píxel obtienen cero.
      Subestimación de -28% a -63%.

    SOLUCIÓN: exactextract (Baston, 2022) calcula la fracción geométrica
    exacta de cada píxel cubierta por cada polígono:
        sum = Σ(xi × ci), donde ci = fracción del píxel i cubierta
    Propiedad verificada: Σ parroquias = total provincial (error < 0.01%).

    Referencia: Baston, D. (2022). exactextract: Fast and accurate raster
    zonal statistics. https://github.com/isciences/exactextract

 CORRECCIONES HEREDADAS DE v3.0:
    - Quinua solo provincial (ESPAC), no desagregable a parroquias
    - CONALI en lugar de GADM v4.1 para límites parroquiales
    - Sin valores hardcodeados (falla explícita si ESPAC no parseable)
    - Comparación MapSPAM vs ESPAC sin ajuste forzado
    - Variables consistentes (Harvested Area en ambas fuentes)

================================================================================
 METODOLOGÍA
================================================================================

 1. DESAGREGACIÓN ESPACIAL (papa, maíz, fréjol):
    - Fuente raster: MapSPAM v2r0 2020 (IFPRI/Harvard Dataverse)
      Variable: Harvested Area (ha/pixel) | Resolución: ~10 km (5 arc-min)
    - Fuente vectorial: CONALI/INEC, límites parroquiales oficiales
    - Método: Estadísticas zonales con fracción de cobertura exacta
      (exactextract). Fórmula: sum = Σ(xi × ci), donde ci es la fracción
      geométrica del píxel i cubierta por el polígono parroquial.
    - Verificación: Propiedad de conservación — la suma de las 42
      parroquias debe ser igual al total provincial como polígono único.
    - Referencia: Baston, D. (2022). exactextract: Fast and accurate
      raster zonal statistics. https://github.com/isciences/exactextract
    - Referencia: You, L., Wood-Sichra, U., Fritz, S., Guo, Z., See, L.,
      & Koo, J. (2014). Spatial Production Allocation Model (SPAM) 2005
      v3.2. IFPRI Discussion Paper 01400. doi:10.7910/DVN/DHXBJX

 2. VALIDACIÓN PROVINCIAL (papa, maíz, fréjol):
    - Total parroquial MapSPAM (Harvested Area) se compara contra total
      provincial ESPAC 2024 (Superficie Cosechada, columna 4).
    - Se reporta ratio MapSPAM/ESPAC como indicador de consistencia.
    - NO se aplica factor de ajuste forzado: MapSPAM y ESPAC son fuentes
      independientes con metodologías distintas; forzar coincidencia
      numérica enmascararía discrepancias reales.
    - Referencia: INEC (2024). ESPAC 2024. Metodología: "La ESPAC no
      permite hacer mayores desagregaciones geográficas (existe
      representatividad hasta nivel provincial)."

 3. QUINUA — LIMITACIÓN DOCUMENTADA:
    - MapSPAM no incluye quinua como cultivo individual.
    - ESPAC proporciona dato provincial (Imbabura) con representatividad
      estadística certificada, pero su diseño muestral NO permite
      desagregación sub-provincial (INEC, 2022, Sección 1.6).
    - No existe fuente espacial con cobertura verificada de quinua a
      nivel parroquial en Imbabura.
    - DECISIÓN: Se reporta quinua como total provincial ESPAC. NO se
      distribuye a parroquias. Esta limitación se documenta explícitamente
      en la tesis como restricción de datos para el componente de exposición.
    - Fuentes agotadas: MapSPAM (no existe), SIGTIERRAS WFS (sin cobertura
      para cultivos andinos de sierra), PDOTs (formato PDF heterogéneo sin
      estandarización, no verificable automáticamente).

================================================================================
 FUENTES DE DATOS
================================================================================

 | Fuente       | Uso                        | Variable        | Resolución  |
 |--------------|----------------------------|-----------------|-------------|
 | MapSPAM v2r0 | Distribución espacial      | Harvested Area  | ~10 km      |
 |              | papa, maíz, fréjol         | (ha/pixel)      |             |
 | ESPAC 2024   | Validación provincial      | Sup. Cosechada  | Provincial  |
 |              | + dato quinua provincial    | (ha)            |             |
 | CONALI/INEC  | Límites parroquiales       | Polígonos ADM3  | Oficial     |

================================================================================
 NORMAS DE CALIDAD
================================================================================

 - ISO 19115:2014 Metadatos geográficos
 - ISO 19157:2013 Calidad de datos geográficos
 - Principios FAIR (Findable, Accessible, Interoperable, Reusable)
 - Trazabilidad completa con auditoría automática

================================================================================
 REFERENCIAS BIBLIOGRÁFICAS
================================================================================

 IFPRI (2024). Spatial Production Allocation Model (SPAM) 2020 V2r0.
     Harvard Dataverse. https://doi.org/10.7910/DVN/SWPENT

 Baston, D. (2022). exactextract: Fast and accurate raster zonal
     statistics. https://github.com/isciences/exactextract
     Método: Fracción de cobertura exacta. sum = Σ(xi × ci), donde
     ci = fracción geométrica del píxel cubierta por el polígono.

 You, L., Wood-Sichra, U., Fritz, S., Guo, Z., See, L., & Koo, J.
     (2014). Spatial Production Allocation Model (SPAM) 2005 v3.2.
     IFPRI Discussion Paper 01400. Washington, DC: IFPRI.

 INEC (2024). Encuesta de Superficie y Producción Agropecuaria Continua
     (ESPAC) 2024. Instituto Nacional de Estadística y Censos del Ecuador.
     www.ecuadorencifras.gob.ec

 INEC (2022). Metodología ESPAC 2022. Sección 1.6 Limitaciones del
     estudio: "La ESPAC, debido al diseño muestral no permite hacer
     mayores desagregaciones geográficas (existe representatividad
     hasta nivel provincial)."

 CONALI (2024). Comité Nacional de Límites Internos. División político-
     administrativa oficial del Ecuador.

 IPCC (2014). Climate Change 2014: Impacts, Adaptation, and Vulnerability.
     Cambridge University Press.

================================================================================
"""

# =============================================================================
# IMPORTACIONES
# =============================================================================
import os
import sys
import json
import time
import zipfile
import hashlib
import warnings
import traceback
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

# Verificar dependencias geoespaciales
DEPENDENCIAS_FALTANTES = []
try:
    import geopandas as gpd
except ImportError:
    DEPENDENCIAS_FALTANTES.append("geopandas")
try:
    import rasterio
    from rasterio.mask import mask as rasterio_mask
except ImportError:
    DEPENDENCIAS_FALTANTES.append("rasterio")
try:
    from exactextract import exact_extract
except ImportError:
    DEPENDENCIAS_FALTANTES.append("exactextract")
try:
    import openpyxl
except ImportError:
    DEPENDENCIAS_FALTANTES.append("openpyxl")
try:
    import requests
except ImportError:
    DEPENDENCIAS_FALTANTES.append("requests")

if DEPENDENCIAS_FALTANTES:
    print("=" * 70)
    print(" ERROR: Dependencias faltantes")
    print("=" * 70)
    print(f"\n Instalar con: pip install {' '.join(DEPENDENCIAS_FALTANTES)}")
    print()
    sys.exit(1)

warnings.filterwarnings('ignore', category=FutureWarning)
warnings.filterwarnings('ignore', category=RuntimeWarning)


# =============================================================================
# CONFIGURACIÓN
# =============================================================================

@dataclass
class ConfiguracionProyecto:
    """Configuración centralizada del proyecto - Script 04C v3.0."""

    # =========================================================================
    # METADATOS
    # =========================================================================
    TITULO_TESIS: str = (
        "Riesgo agroclimático de cultivos andinos bajo escenarios CMIP6 "
        "en la provincia de Imbabura: modelamiento de distribución de "
        "especies para la gestión territorial"
    )
    AUTOR: str = "Víctor Hugo Pinto Páez"
    UNIVERSIDAD: str = "Universidad San Gregorio de Portoviejo"
    PROGRAMA: str = (
        "Maestría en Prevención y Gestión de Riesgos - Mención en "
        "Variabilidad Climática y Resiliencia Territorial"
    )
    SCRIPT_ID: str = "04C"
    SCRIPT_DESC: str = "Desagregación espacial de superficie agrícola por parroquia"
    VERSION: str = "3.0.0"
    TIMESTAMP: str = field(default_factory=lambda: datetime.now().strftime("%Y%m%d_%H%M%S"))
    FECHA: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d"))

    # =========================================================================
    # RUTAS PRINCIPALES
    # =========================================================================
    BASE_DIR: Path = field(default_factory=lambda: Path(
        r"<RUTA_LOCAL>"
    ))

    @property
    def DATA_DIR(self) -> Path:
        return self.BASE_DIR / "02_DATOS"

    @property
    def OUTPUT_DIR(self) -> Path:
        return self.BASE_DIR / "04_RESULTADOS" / "exposicion_agricola"

    @property
    def DOWNLOADS_DIR(self) -> Path:
        return self.BASE_DIR / "02_DATOS" / "descargas_04C"

    @property
    def AUDITORIA_DIR(self) -> Path:
        return self.BASE_DIR / "05_DOCUMENTACION" / "reportes_auditoria"

    @property
    def METADATOS_DIR(self) -> Path:
        return self.BASE_DIR / "05_DOCUMENTACION" / "metadatos_iso"

    # =========================================================================
    # LÍMITES PARROQUIALES — FUENTE OFICIAL CONALI/INEC
    # =========================================================================
    # DECISIÓN METODOLÓGICA: Se usa el archivo oficial del Estado ecuatoriano
    # (CONALI) en lugar de GADM v4.1 (compilación internacional de UC Davis).
    # Justificación: CONALI es la autoridad legal en materia de límites
    # político-administrativos del Ecuador.
    # =========================================================================
    PARROQUIAS_FILE: Path = field(default_factory=lambda: Path(
        r"<RUTA_LOCAL>"
    ))

    # Columnas del GeoPackage CONALI (estructura verificada)
    COL_PARROQUIA: str = "DPA_DESPAR"     # Nombre de parroquia
    COL_CANTON: str = "DPA_DESCAN"        # Nombre de cantón
    COL_CODIGO_PARROQUIA: str = "DPA_PARROQ"  # Código DPA parroquial

    # =========================================================================
    # URLs DE DESCARGA
    # =========================================================================
    MAPSPAM_URL: str = "https://dataverse.harvard.edu/api/access/datafile/11596412"
    ESPAC_URL: str = (
        "https://www.ecuadorencifras.gob.ec/documentos/web-inec/"
        "Estadisticas_agropecuarias/espac/2024/Tabulados_ESPAC_2024.xlsx"
    )

    # =========================================================================
    # MAPSPAM: CORRESPONDENCIA CULTIVO → ARCHIVO
    # =========================================================================
    # Convención MapSPAM: spam2020_V2r0_global_H_{CROP}_{TECH}.tif
    # H = Harvested Area, _A = All technologies
    # NOTA: Quinua NO existe en MapSPAM (limitación documentada)
    # =========================================================================
    MAPSPAM_CROPS: Dict = field(default_factory=lambda: {
        "papa":   {"code": "POTA", "file": "spam2020_V2r0_global_H_POTA_A.tif"},
        "maiz":   {"code": "MAIZ", "file": "spam2020_V2r0_global_H_MAIZ_A.tif"},
        "frejol": {"code": "BEAN", "file": "spam2020_V2r0_global_H_BEAN_A.tif"},
    })
    # Quinua explícitamente EXCLUIDA de MapSPAM — no inventar datos

    # =========================================================================
    # ESPAC 2024: HOJAS CON DATOS POR CULTIVO
    # =========================================================================
    # CORRECCIÓN v3.0: Se usa Superficie COSECHADA (columna 4, 0-indexed)
    # para ser consistente con MapSPAM que reporta Harvested Area.
    # Superficie Plantada ≠ Superficie Cosechada (plantada ≥ cosechada).
    # Mezclar variables distintas introduce sesgo sistemático.
    # =========================================================================
    ESPAC_HOJAS: Dict = field(default_factory=lambda: {
        # Hoja: (cultivo_destino, nombre_cultivo_espac)
        "T43": ("papa",   "Papa (tubérculo fresco)"),
        "T38": ("maiz",   "Maíz duro choclo"),
        "T39": ("maiz",   "Maíz duro seco"),
        "T40": ("maiz",   "Maíz suave choclo"),
        "T41": ("maiz",   "Maíz suave seco"),
        "T34": ("frejol", "Fréjol seco (grano seco)"),
        "T44": ("quinua", "Quinua (grano seco)"),
    })

    # =========================================================================
    # COLUMNAS ESPAC (0-indexed en la estructura del tabulado)
    # Col 3 = Superficie Plantada (ha)
    # Col 4 = Superficie Cosechada (ha)
    # =========================================================================
    ESPAC_COL_COSECHADA: int = 4   # Consistente con MapSPAM Harvested Area
    ESPAC_COL_PLANTADA: int = 3    # Para referencia/reporte


# =============================================================================
# CLASE PRINCIPAL
# =============================================================================

class ExposicionAgricola:
    """
    Pipeline de desagregación espacial de superficie agrícola por parroquia
    para el componente de exposición del modelo de riesgo agroclimático.

    PRINCIPIO RECTOR: Toda decisión metodológica debe tener respaldo en
    literatura científica o en documentación oficial de las fuentes de datos.
    Cuando no existe dato verificable, se documenta la limitación en lugar
    de fabricar estimaciones sin fundamento empírico.
    """

    def __init__(self, config: ConfiguracionProyecto):
        self.config = config
        self.timestamp_inicio = datetime.now()
        self.log = []
        self.errores = []
        self.advertencias = []
        self.decisiones = []
        self.resultados = {}

        # Crear directorios
        for d in [self.config.OUTPUT_DIR, self.config.DOWNLOADS_DIR,
                  self.config.AUDITORIA_DIR, self.config.METADATOS_DIR]:
            d.mkdir(parents=True, exist_ok=True)

        self._log("=" * 70)
        self._log("SCRIPT 04C v3.0: DESAGREGACIÓN ESPACIAL DE EXPOSICIÓN AGRÍCOLA")
        self._log("=" * 70)
        self._log(f"Inicio: {self.timestamp_inicio.strftime('%Y-%m-%d %H:%M:%S')}")
        self._log(f"Versión: {self.config.VERSION}")

    def _log(self, msg: str, level: str = "INFO"):
        """Log con registro para auditoría."""
        ts = datetime.now().strftime('%H:%M:%S')
        entry = f"[{ts}] [{level}] {msg}"
        self.log.append(entry)
        print(entry)
        if level == "ERROR":
            self.errores.append(msg)
        elif level == "WARNING":
            self.advertencias.append(msg)

    def _registrar_decision(self, codigo: str, descripcion: str,
                            justificacion: str, referencia: str):
        """Registra una decisión metodológica con su justificación."""
        decision = {
            "codigo": codigo,
            "descripcion": descripcion,
            "justificacion": justificacion,
            "referencia": referencia,
            "timestamp": datetime.now().isoformat()
        }
        self.decisiones.append(decision)
        self._log(f"  DECISIÓN {codigo}: {descripcion}")
        self._log(f"    Justificación: {justificacion}")
        self._log(f"    Referencia: {referencia}")

    # =========================================================================
    # PASO 1: DESCARGA DE DATOS
    # =========================================================================

    def _descargar_archivo(self, url: str, destino: Path, descripcion: str,
                           max_intentos: int = 3) -> bool:
        """Descarga un archivo con reintentos y verificación de integridad."""
        if destino.exists() and destino.stat().st_size > 0:
            size_mb = destino.stat().st_size / 1e6
            self._log(f"  ✓ {descripcion} ya existe ({size_mb:.1f} MB): {destino.name}")
            return True

        self._log(f"  Descargando {descripcion}...")
        for intento in range(1, max_intentos + 1):
            try:
                response = requests.get(url, stream=True, timeout=300,
                                       allow_redirects=True)
                response.raise_for_status()

                total = int(response.headers.get('content-length', 0))
                descargado = 0

                with open(destino, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192 * 16):
                        f.write(chunk)
                        descargado += len(chunk)
                        if total > 0 and descargado % (1024 * 1024 * 10) < 8192 * 16:
                            pct = (descargado / total) * 100
                            self._log(f"    Progreso: {pct:.0f}% "
                                     f"({descargado/1e6:.0f}/{total/1e6:.0f} MB)")

                size_mb = destino.stat().st_size / 1e6
                md5 = hashlib.md5()
                with open(destino, 'rb') as f:
                    for chunk in iter(lambda: f.read(8192 * 16), b''):
                        md5.update(chunk)
                self._log(f"  ✓ {descripcion}: {size_mb:.1f} MB | MD5: {md5.hexdigest()}")
                return True

            except Exception as e:
                self._log(f"  ✗ Intento {intento}/{max_intentos}: {e}", "WARNING")
                if intento < max_intentos:
                    time.sleep(5 * intento)

        self._log(f"  ✗ FALLO tras {max_intentos} intentos: {descripcion}", "ERROR")
        return False

    def descargar_datos(self) -> Dict[str, bool]:
        """Paso 1: Descarga fuentes de datos necesarias."""
        self._log("\n" + "=" * 70)
        self._log("[1/5] DESCARGA DE DATOS FUENTE")
        self._log("=" * 70)

        resultados = {}

        # 1.1 MapSPAM 2020 (área cosechada — papa, maíz, fréjol)
        mapspam_zip = self.config.DOWNLOADS_DIR / "mapspam_harvested_area.zip"
        resultados['mapspam'] = self._descargar_archivo(
            self.config.MAPSPAM_URL, mapspam_zip,
            "MapSPAM 2020 V2r0 Harvested Area (IFPRI)"
        )

        # 1.2 Extraer rasters específicos del ZIP
        if resultados['mapspam']:
            self._log("  Extrayendo rasters de cultivos objetivo...")
            mapspam_dir = self.config.DOWNLOADS_DIR / "mapspam_geotiff"
            mapspam_dir.mkdir(exist_ok=True)

            try:
                with zipfile.ZipFile(mapspam_zip, 'r') as zf:
                    all_names = zf.namelist()
                    for cultivo, info in self.config.MAPSPAM_CROPS.items():
                        output_path = mapspam_dir / info['file']
                        if output_path.exists() and output_path.stat().st_size > 0:
                            self._log(f"    ✓ {cultivo}: ya extraído")
                            continue

                        # Buscar el archivo dentro del ZIP
                        target = None
                        for name in all_names:
                            if info['file'] in name:
                                target = name
                                break

                        if target:
                            zf.extract(target, self.config.DOWNLOADS_DIR)
                            extracted = self.config.DOWNLOADS_DIR / target
                            if extracted != output_path:
                                extracted.rename(output_path)
                            self._log(f"    ✓ {cultivo}: {info['file']}")
                        else:
                            self._log(f"    ✗ {cultivo}: no encontrado en ZIP", "ERROR")
                            resultados['mapspam'] = False
            except zipfile.BadZipFile:
                self._log("    ✗ Archivo ZIP corrupto, eliminar y re-descargar", "ERROR")
                resultados['mapspam'] = False

            # Limpiar directorio temporal
            for subdir in self.config.DOWNLOADS_DIR.iterdir():
                if subdir.is_dir() and subdir.name.startswith("spam"):
                    import shutil
                    shutil.rmtree(subdir, ignore_errors=True)

        # 1.3 ESPAC 2024 (validación provincial + dato quinua)
        espac_file = self.config.DOWNLOADS_DIR / "Tabulados_ESPAC_2024.xlsx"
        resultados['espac'] = self._descargar_archivo(
            self.config.ESPAC_URL, espac_file, "ESPAC 2024 Tabulados (INEC)"
        )

        # 1.4 Verificar archivo de parroquias CONALI
        if self.config.PARROQUIAS_FILE.exists():
            self._log(f"  ✓ Parroquias CONALI: {self.config.PARROQUIAS_FILE.name}")
            resultados['parroquias'] = True
        else:
            self._log(f"  ✗ Archivo de parroquias no encontrado: "
                     f"{self.config.PARROQUIAS_FILE}", "ERROR")
            resultados['parroquias'] = False

        self.resultados['descargas'] = resultados
        return resultados

    # =========================================================================
    # PASO 2: CARGAR LÍMITES PARROQUIALES (CONALI)
    # =========================================================================

    def cargar_parroquias(self) -> gpd.GeoDataFrame:
        """
        Paso 2: Carga parroquias de Imbabura desde archivo oficial CONALI.

        DECISIÓN METODOLÓGICA D1: Se usa CONALI (Comité Nacional de Límites
        Internos) en lugar de GADM v4.1 porque CONALI es la fuente oficial
        del Estado ecuatoriano para división político-administrativa.
        """
        self._log("\n" + "=" * 70)
        self._log("[2/5] CARGA DE LÍMITES PARROQUIALES (CONALI)")
        self._log("=" * 70)

        self._registrar_decision(
            "D1",
            "Límites parroquiales: CONALI/INEC en lugar de GADM v4.1",
            "CONALI es la autoridad legal del Estado ecuatoriano para "
            "límites político-administrativos. GADM es una compilación "
            "internacional que puede no coincidir con los límites legales.",
            "CONALI - Comité Nacional de Límites Internos (Ecuador)"
        )

        imbabura = gpd.read_file(self.config.PARROQUIAS_FILE)
        self._log(f"  Parroquias cargadas: {len(imbabura)}")
        self._log(f"  CRS: {imbabura.crs}")
        self._log(f"  Columnas: {imbabura.columns.tolist()}")

        # Mapear columnas CONALI a nombres internos estándar
        # Estructura CONALI verificada: DPA_DESPAR (parroquia), DPA_DESCAN (cantón),
        # DPA_PARROQ (código DPA), DPA_PROVIN (código provincia)
        col_parroquia = self.config.COL_PARROQUIA
        col_canton = self.config.COL_CANTON
        col_codigo = self.config.COL_CODIGO_PARROQUIA

        # Verificar que las columnas existan
        for col_name, col_val in [("parroquia", col_parroquia),
                                   ("cantón", col_canton),
                                   ("código DPA", col_codigo)]:
            if col_val not in imbabura.columns:
                self._log(f"  ✗ Columna '{col_val}' ({col_name}) no encontrada "
                         f"en {self.config.PARROQUIAS_FILE.name}", "ERROR")
                self._log(f"    Columnas disponibles: {imbabura.columns.tolist()}")
                raise KeyError(
                    f"Columna '{col_val}' no encontrada. Verifique "
                    f"ConfiguracionProyecto.COL_PARROQUIA/COL_CANTON"
                )

        self._log(f"  Columna cantón: {col_canton}")
        self._log(f"  Columna parroquia: {col_parroquia}")
        self._log(f"  Columna código DPA: {col_codigo}")

        # Renombrar para consistencia interna del pipeline
        imbabura = imbabura.rename(columns={
            col_parroquia: 'parroquia',
            col_canton: 'canton',
            col_codigo: 'codigo_dpa',
        })

        # Asegurar proyección UTM para cálculo de áreas
        if imbabura.crs and imbabura.crs.to_epsg() == 32717:
            imbabura['area_km2'] = imbabura.geometry.area / 1e6
        else:
            imbabura_utm = imbabura.to_crs(epsg=32717)
            imbabura['area_km2'] = imbabura_utm.geometry.area / 1e6

        # Para operaciones con rasters (WGS84), crear versión reproyectada
        if imbabura.crs and imbabura.crs.to_epsg() != 4326:
            self.parroquias_wgs84 = imbabura.to_crs(epsg=4326)
        else:
            self.parroquias_wgs84 = imbabura.copy()

        # Listar parroquias
        self._log(f"\n  {'CANTÓN':<20s} {'PARROQUIA':<30s} {'ÁREA (km²)':>12s}")
        self._log(f"  {'-'*20} {'-'*30} {'-'*12}")
        for _, row in imbabura.sort_values(
                ['canton', 'parroquia'] if 'canton' in imbabura.columns
                else ['parroquia']).iterrows():
            canton = row.get('canton', 'N/A')
            parroquia = row.get('parroquia', 'N/A')
            area = row['area_km2']
            self._log(f"  {str(canton):<20s} {str(parroquia):<30s} {area:>10.1f}")

        self._log(f"\n  Total: {len(imbabura)} parroquias | "
                  f"Área total: {imbabura['area_km2'].sum():.0f} km²")

        self.parroquias = imbabura
        self.resultados['n_parroquias'] = len(imbabura)
        return imbabura

    # =========================================================================
    # PASO 3: ESTADÍSTICAS ZONALES MapSPAM (PAPA, MAÍZ, FRÉJOL)
    # =========================================================================

    def calcular_zonal_stats(self) -> pd.DataFrame:
        """
        Paso 3: Estadísticas zonales MapSPAM × parroquias.

        SOLO para papa, maíz, fréjol (cultivos disponibles en MapSPAM).
        Quinua NO se procesa aquí — ver documentación de limitación.

        MÉTODO: exactextract con fracción de cobertura exacta.
        Fórmula: sum = Σ(xi × ci), donde ci = fracción geométrica del
        píxel i cubierta por el polígono parroquial.

        VERIFICACIÓN: Propiedad de conservación — la suma de las 42
        parroquias debe ser igual al total provincial como polígono único.

        Referencia: Baston, D. (2022). exactextract.
        """
        self._log("\n" + "=" * 70)
        self._log("[3/5] ESTADÍSTICAS ZONALES: MapSPAM × PARROQUIAS (exactextract)")
        self._log("=" * 70)

        self._registrar_decision(
            "D2",
            "exactextract con fracción de cobertura para estadísticas zonales",
            "La resolución de MapSPAM (~86 km²/píxel) es mayor que muchas "
            "parroquias de Imbabura (10-80 km²). rasterstats no soporta "
            "píxeles parciales: all_touched=True genera doble conteo "
            "(+215% a +387%), all_touched=False pierde parroquias pequeñas "
            "(-28% a -63%). exactextract calcula la fracción geométrica "
            "exacta: sum = Σ(xi × ci). Propiedad verificada: "
            "Σ_parroquias = total_provincial (error < 0.01%).",
            "Baston, D. (2022). exactextract. "
            "You et al. (2014). SPAM 2005 v3.2. IFPRI."
        )

        mapspam_dir = self.config.DOWNLOADS_DIR / "mapspam_geotiff"
        resultados_cultivos = []

        # Polígono provincial para verificación de conservación
        provincia_union = self.parroquias_wgs84.union_all()
        gdf_provincia = gpd.GeoDataFrame(
            {'nombre': ['IMBABURA']},
            geometry=[provincia_union],
            crs="EPSG:4326"
        )

        for cultivo, info in self.config.MAPSPAM_CROPS.items():
            raster_path = mapspam_dir / info['file']
            self._log(f"\n  Procesando {cultivo.upper()} ({info['code']})...")

            if not raster_path.exists():
                self._log(f"    ✗ Raster no encontrado: {raster_path}", "ERROR")
                continue

            # Verificar metadata del raster
            with rasterio.open(raster_path) as src:
                self._log(f"    Raster: {src.width}×{src.height} | "
                         f"CRS: {src.crs} | NoData: {src.nodata}")
                self._log(f"    Bounds: {src.bounds}")
                res_deg = src.res[0]
                res_km = res_deg * 111.32
                pixel_area_km2 = (res_deg * 111.32) ** 2
                self._log(f"    Resolución: {res_deg:.4f}° ≈ {res_km:.1f} km "
                         f"({pixel_area_km2:.0f} km²/píxel)")

            # ─── exactextract: fracción de cobertura exacta ─────────
            # sum = Σ(xi × ci), ci = fracción geométrica cubierta
            df_parroquial = exact_extract(
                str(raster_path),
                self.parroquias_wgs84,
                ['sum', 'count'],
                include_cols=['parroquia', 'canton'],
                output='pandas'
            )

            # Verificación de conservación: provincia como polígono único
            df_provincial = exact_extract(
                str(raster_path),
                gdf_provincia,
                ['sum'],
                output='pandas'
            )
            total_provincial = df_provincial['sum'].iloc[0]
            total_parroquias = df_parroquial['sum'].sum()

            # Criterio: error < 1%
            if total_provincial > 0:
                error_conservacion = abs(total_parroquias - total_provincial) / total_provincial * 100
            else:
                error_conservacion = 0.0

            self._log(f"    Σ parroquias: {total_parroquias:.1f} ha")
            self._log(f"    Total provincial: {total_provincial:.1f} ha")
            self._log(f"    Error conservación: {error_conservacion:.4f}%")
            if error_conservacion < 1.0:
                self._log(f"    ✓ Conservación verificada (< 1%)")
            else:
                self._log(f"    ⚠ Error de conservación > 1%", "WARNING")

            # Compilar resultados
            for _, row in df_parroquial.iterrows():
                ha_total = max(row['sum'], 0) if not np.isnan(row['sum']) else 0.0
                coverage = row['count'] if not np.isnan(row['count']) else 0.0

                resultados_cultivos.append({
                    'cultivo': cultivo,
                    'canton': str(row['canton']),
                    'parroquia': str(row['parroquia']),
                    'ha_cosechada_mapspam': round(ha_total, 2),
                    'cobertura_pixeles': round(coverage, 2),
                    'area_km2': round(
                        self.parroquias.loc[
                            self.parroquias['parroquia'] == row['parroquia'],
                            'area_km2'
                        ].iloc[0], 1
                    ) if row['parroquia'] in self.parroquias['parroquia'].values else 0,
                    'fuente': 'MapSPAM v2r0 2020',
                    'variable': 'Harvested Area (ha)',
                    'metodo': 'exactextract sum=Σ(xi×ci)',
                    'nivel_dato': 'parroquial',
                })

            # Resumen por cultivo
            n_con_cultivo = sum(1 for r in resultados_cultivos
                               if r['cultivo'] == cultivo and r['ha_cosechada_mapspam'] > 0)
            self._log(f"    ✓ Total Imbabura: {total_parroquias:.1f} ha en "
                     f"{n_con_cultivo}/{len(self.parroquias)} parroquias")

        df = pd.DataFrame(resultados_cultivos)
        self.df_mapspam = df
        return df

    # =========================================================================
    # PASO 4: EXTRACCIÓN ESPAC (VALIDACIÓN + DATO QUINUA PROVINCIAL)
    # =========================================================================

    def parsear_espac(self) -> Dict:
        """
        Paso 4: Extraer datos provinciales de ESPAC 2024.

        Se extraen DOS variables para cada cultivo:
        - Superficie Plantada (ha) — columna 3
        - Superficie Cosechada (ha) — columna 4

        Para validación contra MapSPAM se usa Cosechada (consistencia de
        variable). Quinua se reporta como dato provincial ÚNICAMENTE.

        Si no se puede parsear algún cultivo, el script FALLA.
        No se usan valores hardcodeados.
        """
        self._log("\n" + "=" * 70)
        self._log("[4/5] EXTRACCIÓN DE DATOS PROVINCIALES (ESPAC 2024)")
        self._log("=" * 70)

        self._registrar_decision(
            "D3",
            "ESPAC como fuente de validación provincial, NO para desagregación",
            "La metodología ESPAC (INEC, 2022, Sección 1.6) establece "
            "explícitamente: 'La ESPAC, debido al diseño muestral no permite "
            "hacer mayores desagregaciones geográficas (existe representatividad "
            "hasta nivel provincial)'. Usar datos ESPAC a nivel sub-provincial "
            "viola las condiciones de representatividad estadística del diseño "
            "muestral.",
            "INEC (2022). Metodología ESPAC 2022. Sección 1.6."
        )

        espac_file = self.config.DOWNLOADS_DIR / "Tabulados_ESPAC_2024.xlsx"

        if not espac_file.exists():
            self._log("  ✗ ESPAC no encontrado.", "ERROR")
            raise FileNotFoundError(
                f"Archivo ESPAC requerido: {espac_file}\n"
                "Descargue de: https://www.ecuadorencifras.gob.ec/"
                "documentos/web-inec/Estadisticas_agropecuarias/espac/2024/"
            )

        self._log(f"  Leyendo {espac_file.name}...")
        wb = openpyxl.load_workbook(espac_file, read_only=True, data_only=True)
        self._log(f"  Hojas disponibles: {len(wb.sheetnames)}")

        # Almacenamiento
        espac_cosechada = {}    # cultivo → ha cosechada
        espac_plantada = {}     # cultivo → ha plantada
        espac_detalle = {}      # hoja → (nombre, ha_plantada, ha_cosechada)

        col_plantada = self.config.ESPAC_COL_PLANTADA
        col_cosechada = self.config.ESPAC_COL_COSECHADA

        for hoja, (cultivo_dest, nombre_espac) in self.config.ESPAC_HOJAS.items():
            if hoja not in wb.sheetnames:
                self._log(f"  ✗ Hoja '{hoja}' no encontrada", "ERROR")
                raise KeyError(f"Hoja '{hoja}' no en {espac_file.name}")

            ws = wb[hoja]
            val_plantada = None
            val_cosechada = None

            for row in ws.iter_rows(values_only=True):
                row_vals = list(row)
                # Buscar fila "IMBABURA"
                if (len(row_vals) > max(col_plantada, col_cosechada) and
                    row_vals[1] is not None and
                    str(row_vals[1]).strip().upper() == "IMBABURA"):

                    v_plant = row_vals[col_plantada]
                    v_cosec = row_vals[col_cosechada]

                    if isinstance(v_plant, (int, float)) and v_plant >= 0:
                        val_plantada = float(v_plant)
                    if isinstance(v_cosec, (int, float)) and v_cosec >= 0:
                        val_cosechada = float(v_cosec)
                    break

            if val_plantada is None and val_cosechada is None:
                self._log(f"  ✗ No se encontró Imbabura en '{hoja}' ({nombre_espac})",
                         "ERROR")
                raise ValueError(
                    f"No se encontró 'IMBABURA' con valores numéricos en "
                    f"hoja '{hoja}' ({nombre_espac})"
                )

            # Acumular (maíz tiene 4 hojas)
            for d, val in [(espac_plantada, val_plantada),
                           (espac_cosechada, val_cosechada)]:
                if val is not None:
                    if cultivo_dest not in d:
                        d[cultivo_dest] = 0.0
                    d[cultivo_dest] += val

            espac_detalle[hoja] = (nombre_espac,
                                   val_plantada if val_plantada else 0,
                                   val_cosechada if val_cosechada else 0)

            self._log(f"    {hoja} {nombre_espac:<30s}: "
                     f"Plant={val_plantada:>8.1f} ha | "
                     f"Cosech={val_cosechada:>8.1f} ha")

        wb.close()

        # Verificar cultivos
        cultivos_requeridos = {"papa", "maiz", "frejol", "quinua"}
        faltantes_c = cultivos_requeridos - set(espac_cosechada.keys())
        faltantes_p = cultivos_requeridos - set(espac_plantada.keys())
        if faltantes_c or faltantes_p:
            self._log(f"  ✗ Faltan datos: cosechada={faltantes_c}, "
                     f"plantada={faltantes_p}", "ERROR")
            raise ValueError(f"Datos ESPAC incompletos")

        # Reporte
        self._log(f"\n  Totales provinciales ESPAC 2024 — Imbabura:")
        self._log(f"  {'Cultivo':<10s} {'Plantada (ha)':>14s} {'Cosechada (ha)':>15s}")
        self._log(f"  {'-'*10} {'-'*14} {'-'*15}")
        for c in ['papa', 'maiz', 'frejol', 'quinua']:
            nota = " *" if c == 'quinua' else ""
            self._log(f"  {c:<10s} {espac_plantada[c]:>14.2f} "
                     f"{espac_cosechada[c]:>15.2f}{nota}")
        self._log(f"\n  * Quinua: dato provincial ÚNICAMENTE. No se desagrega "
                 f"a parroquias.")
        self._log(f"    Justificación: ESPAC no tiene representatividad "
                 f"sub-provincial (INEC, 2022).")

        self.espac_cosechada = espac_cosechada
        self.espac_plantada = espac_plantada
        self.espac_detalle = espac_detalle
        return {'cosechada': espac_cosechada, 'plantada': espac_plantada}

    # =========================================================================
    # PASO 5: INTEGRACIÓN, VALIDACIÓN Y PRODUCTOS
    # =========================================================================

    def integrar_y_validar(self) -> pd.DataFrame:
        """
        Paso 5: Integrar resultados, validar, generar productos.

        CAMBIO CRÍTICO v3.0: NO se aplica factor de ajuste forzado.
        Se reporta la discrepancia MapSPAM vs ESPAC como métrica de
        consistencia, pero no se fuerza coincidencia numérica.
        """
        self._log("\n" + "=" * 70)
        self._log("[5/5] INTEGRACIÓN, VALIDACIÓN Y PRODUCTOS FINALES")
        self._log("=" * 70)

        self._registrar_decision(
            "D4",
            "Comparación MapSPAM vs ESPAC sin ajuste forzado",
            "MapSPAM y ESPAC son fuentes independientes con metodologías "
            "distintas. MapSPAM usa modelamiento de asignación espacial; "
            "ESPAC usa encuesta muestral. Forzar coincidencia numérica "
            "mediante un factor lineal enmascara discrepancias reales entre "
            "fuentes y no tiene respaldo metodológico.",
            "You et al. (2014). IFPRI. / INEC (2024). ESPAC Metodología."
        )

        self._registrar_decision(
            "D5",
            "Quinua NO se desagrega a nivel parroquial",
            "No existe fuente verificada de distribución espacial de quinua "
            "a nivel parroquial en Imbabura. MapSPAM no incluye quinua. "
            "ESPAC tiene representatividad solo provincial (INEC, 2022). "
            "Distribuir proporcionalmente por aptitud SDM, área u otro proxy "
            "no tiene respaldo en literatura. Se documenta como limitación.",
            "INEC (2022). Metodología ESPAC, Sección 1.6."
        )

        # -----------------------------------------------------------------
        # 5.1 COMPARACIÓN MapSPAM vs ESPAC (papa, maíz, fréjol)
        # -----------------------------------------------------------------
        self._log("\n  5.1 COMPARACIÓN MapSPAM vs ESPAC (Superficie Cosechada)")
        self._log(f"  {'Cultivo':<10s} {'MapSPAM (ha)':>14s} {'ESPAC (ha)':>12s} "
                 f"{'Ratio':>8s} {'Discrepancia':>14s}")
        self._log(f"  {'-'*10} {'-'*14} {'-'*12} {'-'*8} {'-'*14}")

        comparacion = {}
        for cultivo in ['papa', 'maiz', 'frejol']:
            mask = self.df_mapspam['cultivo'] == cultivo
            total_mapspam = self.df_mapspam.loc[mask, 'ha_cosechada_mapspam'].sum()
            total_espac = self.espac_cosechada.get(cultivo, 0)

            if total_espac > 0:
                ratio = total_mapspam / total_espac
            else:
                ratio = float('inf')

            discrepancia = total_mapspam - total_espac
            self._log(f"  {cultivo:<10s} {total_mapspam:>14.1f} {total_espac:>12.1f} "
                     f"{ratio:>8.3f} {discrepancia:>+14.1f}")

            comparacion[cultivo] = {
                'mapspam_ha': round(total_mapspam, 2),
                'espac_cosechada_ha': round(total_espac, 2),
                'ratio_mapspam_espac': round(ratio, 4),
                'discrepancia_ha': round(discrepancia, 2),
            }

        # Quinua: solo dato provincial
        self._log(f"\n  Quinua (dato provincial ESPAC únicamente):")
        self._log(f"    Plantada: {self.espac_plantada['quinua']:.2f} ha")
        self._log(f"    Cosechada: {self.espac_cosechada['quinua']:.2f} ha")
        self._log(f"    Nivel: PROVINCIAL (no desagregable)")

        # -----------------------------------------------------------------
        # 5.2 TABLA DE EXPOSICIÓN PARROQUIAL (papa, maíz, fréjol)
        # -----------------------------------------------------------------
        self._log("\n  5.2 TABLA DE EXPOSICIÓN PARROQUIAL")

        df_parroquial = self.df_mapspam.copy()

        # Tabla pivote: parroquia × cultivo
        pivot = df_parroquial.pivot_table(
            index=['canton', 'parroquia', 'area_km2'],
            columns='cultivo',
            values='ha_cosechada_mapspam',
            aggfunc='sum',
            fill_value=0
        ).reset_index()

        # Total agrícola (sin quinua — no hay dato parroquial)
        cultivos_cols = [c for c in ['papa', 'maiz', 'frejol']
                        if c in pivot.columns]
        pivot['total_ha_sin_quinua'] = pivot[cultivos_cols].sum(axis=1)
        pivot['densidad_agricola_ha_km2'] = (
            pivot['total_ha_sin_quinua'] / pivot['area_km2']
        ).round(4)
        pivot['n_cultivos'] = (pivot[cultivos_cols] > 0).sum(axis=1)

        # Ordenar por superficie total
        pivot = pivot.sort_values('total_ha_sin_quinua', ascending=False)

        # Mostrar
        self._log(f"\n  {'Cantón':<18s} {'Parroquia':<25s} {'Papa':>8s} "
                 f"{'Maíz':>8s} {'Fréjol':>8s} {'TOTAL':>8s}")
        self._log(f"  {'-'*18} {'-'*25} {'-'*8} {'-'*8} {'-'*8} {'-'*8}")
        for _, row in pivot.head(15).iterrows():
            self._log(f"  {str(row['canton']):<18s} {str(row['parroquia']):<25s} "
                     f"{row.get('papa', 0):>8.1f} {row.get('maiz', 0):>8.1f} "
                     f"{row.get('frejol', 0):>8.1f} "
                     f"{row['total_ha_sin_quinua']:>8.1f}")

        # -----------------------------------------------------------------
        # 5.3 GUARDAR PRODUCTOS
        # -----------------------------------------------------------------
        self._log("\n  5.3 GUARDANDO PRODUCTOS")

        # CSV detallado parroquial (papa, maíz, fréjol)
        csv_parroquial = (self.config.OUTPUT_DIR /
                          f"exposicion_parroquial_{self.config.TIMESTAMP}.csv")
        df_parroquial.to_csv(csv_parroquial, index=False, encoding='utf-8-sig')
        self._log(f"  ✓ {csv_parroquial.name}")

        # CSV resumen pivote
        csv_resumen = (self.config.OUTPUT_DIR /
                       f"exposicion_resumen_{self.config.TIMESTAMP}.csv")
        pivot.to_csv(csv_resumen, index=False, encoding='utf-8-sig')
        self._log(f"  ✓ {csv_resumen.name}")

        # CSV dato quinua provincial (explícitamente separado)
        quinua_data = {
            'cultivo': ['quinua'],
            'nivel': ['provincial'],
            'provincia': ['Imbabura'],
            'ha_plantada_espac': [self.espac_plantada['quinua']],
            'ha_cosechada_espac': [self.espac_cosechada['quinua']],
            'fuente': ['ESPAC 2024 (INEC)'],
            'nota': ['Representatividad solo provincial. No desagregable '
                     'a parroquias según diseño muestral ESPAC (INEC, 2022).'],
        }
        df_quinua = pd.DataFrame(quinua_data)
        csv_quinua = (self.config.OUTPUT_DIR /
                      f"quinua_provincial_{self.config.TIMESTAMP}.csv")
        df_quinua.to_csv(csv_quinua, index=False, encoding='utf-8-sig')
        self._log(f"  ✓ {csv_quinua.name}")

        # GeoPackage con geometrías
        gdf_resultado = self.parroquias.merge(
            pivot, on=['canton', 'parroquia'], how='left',
            suffixes=('', '_pivot')
        )
        cols_keep = ['canton', 'parroquia', 'geometry', 'area_km2']
        cols_keep += cultivos_cols
        cols_keep += ['total_ha_sin_quinua', 'densidad_agricola_ha_km2', 'n_cultivos']
        cols_keep = [c for c in cols_keep if c in gdf_resultado.columns]

        gpkg_output = (self.config.OUTPUT_DIR /
                       f"exposicion_imbabura_{self.config.TIMESTAMP}.gpkg")
        gdf_resultado[cols_keep].to_file(gpkg_output, driver="GPKG")
        self._log(f"  ✓ {gpkg_output.name}")

        # JSON comparación de fuentes (trazabilidad)
        json_comp = (self.config.OUTPUT_DIR /
                     f"comparacion_fuentes_{self.config.TIMESTAMP}.json")
        with open(json_comp, 'w', encoding='utf-8') as f:
            json.dump({
                'descripcion': 'Comparación MapSPAM vs ESPAC — sin ajuste forzado',
                'justificacion': (
                    'MapSPAM y ESPAC son fuentes independientes. '
                    'MapSPAM usa modelo de asignación espacial (You et al., 2014). '
                    'ESPAC usa encuesta muestral con representatividad provincial '
                    '(INEC, 2022). La discrepancia es informativa, no un error '
                    'que requiera corrección forzada.'
                ),
                'variable_comparada': 'Superficie Cosechada (ha)',
                'comparacion': comparacion,
                'quinua': {
                    'nivel': 'provincial',
                    'plantada_ha': self.espac_plantada['quinua'],
                    'cosechada_ha': self.espac_cosechada['quinua'],
                    'limitacion': (
                        'No existe fuente espacial de quinua a nivel parroquial. '
                        'ESPAC solo tiene representatividad provincial. '
                        'MapSPAM no incluye quinua como cultivo individual.'
                    ),
                },
                'espac_detalle': {
                    k: {"nombre": v[0],
                        "plantada_ha": round(v[1], 2),
                        "cosechada_ha": round(v[2], 2)}
                    for k, v in self.espac_detalle.items()
                },
                'timestamp': self.config.TIMESTAMP,
            }, f, indent=2, ensure_ascii=False)
        self._log(f"  ✓ {json_comp.name}")

        self.df_final = df_parroquial
        self.pivot = pivot
        self.comparacion = comparacion
        return df_parroquial

    # =========================================================================
    # REPORTE DE AUDITORÍA
    # =========================================================================

    def generar_reporte_auditoria(self):
        """Genera reporte de auditoría con trazabilidad completa."""
        self._log("\n  GENERANDO REPORTE DE AUDITORÍA")

        timestamp_fin = datetime.now()
        duracion = (timestamp_fin - self.timestamp_inicio).total_seconds()

        r = []
        r.append("=" * 78)
        r.append("  REPORTE DE AUDITORÍA — SCRIPT 04C v3.0")
        r.append("  DESAGREGACIÓN ESPACIAL DE EXPOSICIÓN AGRÍCOLA POR PARROQUIA")
        r.append("=" * 78)
        r.append(f"")
        r.append(f"  Tesis:       {self.config.TITULO_TESIS}")
        r.append(f"  Autor:       {self.config.AUTOR}")
        r.append(f"  Universidad: {self.config.UNIVERSIDAD}")
        r.append(f"  Versión:     {self.config.VERSION}")
        r.append(f"  Inicio:      {self.timestamp_inicio.strftime('%Y-%m-%d %H:%M:%S')}")
        r.append(f"  Fin:         {timestamp_fin.strftime('%Y-%m-%d %H:%M:%S')}")
        r.append(f"  Duración:    {duracion:.0f} segundos")
        r.append(f"")

        # Fuentes
        r.append("  1. FUENTES DE DATOS")
        r.append("  " + "-" * 72)
        r.append(f"  Parroquias:  CONALI/INEC (fuente oficial Estado ecuatoriano)")
        r.append(f"               Archivo: {self.config.PARROQUIAS_FILE.name}")
        r.append(f"  MapSPAM:     v2r0 2020, Harvested Area, All technologies")
        r.append(f"               DOI: 10.7910/DVN/SWPENT (Harvard Dataverse)")
        r.append(f"               Ref: You et al. (2014). IFPRI Discussion Paper 01400")
        r.append(f"  ESPAC:       Tabulados 2024 (INEC Ecuador)")
        r.append(f"               Ref: INEC (2024). ESPAC 2024")
        r.append(f"")

        # Resultados
        r.append("  2. RESULTADOS POR CULTIVO")
        r.append("  " + "-" * 72)
        r.append(f"  {'Cultivo':<10s} {'MapSPAM':>10s} {'ESPAC Cos.':>12s} "
                 f"{'Ratio':>8s} {'Nivel dato':>15s}")
        r.append(f"  {'-'*10} {'-'*10} {'-'*12} {'-'*8} {'-'*15}")
        for cultivo in ['papa', 'maiz', 'frejol']:
            c = self.comparacion.get(cultivo, {})
            r.append(f"  {cultivo:<10s} {c.get('mapspam_ha', 0):>10.1f} "
                     f"{c.get('espac_cosechada_ha', 0):>12.1f} "
                     f"{c.get('ratio_mapspam_espac', 0):>8.3f} "
                     f"{'Parroquial':>15s}")
        r.append(f"  {'quinua':<10s} {'N/A':>10s} "
                 f"{self.espac_cosechada.get('quinua', 0):>12.1f} "
                 f"{'N/A':>8s} {'PROVINCIAL':>15s}")
        r.append(f"")

        # Detalle ESPAC
        if hasattr(self, 'espac_detalle'):
            r.append("  2b. DETALLE ESPAC POR HOJA")
            r.append("  " + "-" * 72)
            r.append(f"  {'Hoja':<6s} {'Cultivo':<30s} {'Plantada':>10s} {'Cosechada':>10s}")
            r.append(f"  {'-'*6} {'-'*30} {'-'*10} {'-'*10}")
            for hoja, (nombre, ha_p, ha_c) in self.espac_detalle.items():
                r.append(f"  {hoja:<6s} {nombre:<30s} {ha_p:>10.1f} {ha_c:>10.1f}")
            r.append(f"")

        # Decisiones metodológicas
        r.append("  3. DECISIONES METODOLÓGICAS")
        r.append("  " + "-" * 72)
        for d in self.decisiones:
            r.append(f"")
            r.append(f"  {d['codigo']}: {d['descripcion']}")
            r.append(f"      Justificación: {d['justificacion']}")
            r.append(f"      Referencia: {d['referencia']}")
        r.append(f"")

        # Limitaciones
        r.append("  4. LIMITACIONES DOCUMENTADAS")
        r.append("  " + "-" * 72)
        r.append("  L1: QUINUA — No existe fuente espacial a nivel parroquial.")
        r.append("      MapSPAM no incluye quinua. ESPAC solo provincial.")
        r.append("      PDOTs parroquiales tienen datos heterogéneos en PDF.")
        r.append("      Impacto: El componente de exposición para quinua se")
        r.append("      reporta SOLO a nivel provincial en el modelo de riesgo.")
        r.append(f"")
        r.append("  L2: RESOLUCIÓN MAPSPAM — ~10 km puede asignar superficie")
        r.append("      a parroquias adyacentes. Mitigación: exactextract con")
        r.append("      fracción de cobertura exacta + verificación de conservación.")
        r.append(f"")
        r.append("  L3: TEMPORALIDAD — MapSPAM 2020, ESPAC 2024. Cuatro años")
        r.append("      de diferencia entre fuentes. Aceptable para análisis de")
        r.append("      distribución espacial relativa, no para valores absolutos.")
        r.append(f"")

        # Criterios de verificación
        r.append("  5. VERIFICACIÓN DE CRITERIOS")
        r.append("  " + "-" * 72)
        criterios = [
            ("Parroquias procesadas (≥40)",
             hasattr(self, 'parroquias') and len(self.parroquias) >= 40),
            ("3 cultivos con datos parroquiales (MapSPAM)",
             hasattr(self, 'df_final') and self.df_final['cultivo'].nunique() == 3),
            ("Quinua documentada como limitación (provincial)",
             hasattr(self, 'espac_cosechada') and 'quinua' in self.espac_cosechada),
            ("Comparación MapSPAM vs ESPAC generada",
             hasattr(self, 'comparacion') and len(self.comparacion) == 3),
            ("GeoPackage de exposición generado",
             any(self.config.OUTPUT_DIR.glob("exposicion_imbabura_*.gpkg"))),
            ("0 errores críticos",
             len(self.errores) == 0),
            ("Todas las decisiones tienen referencia bibliográfica",
             all('referencia' in d for d in self.decisiones)),
        ]
        for desc, cumple in criterios:
            estado = "✓ SÍ" if cumple else "✗ NO"
            r.append(f"  {estado}  {desc}")

        aprobado = all(c[1] for c in criterios)
        r.append(f"")
        r.append(f"  ESTADO: {'APROBADO' if aprobado else 'REQUIERE REVISIÓN'}")
        r.append(f"")
        r.append(f"  Advertencias: {len(self.advertencias)}")
        r.append(f"  Errores: {len(self.errores)}")
        r.append(f"")

        # Guardar
        reporte_path = (self.config.AUDITORIA_DIR /
                       f"REPORTE_04C_v3_{self.config.TIMESTAMP}.txt")
        with open(reporte_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(r))
        self._log(f"  ✓ Reporte: {reporte_path.name}")

        reporte_output = (self.config.OUTPUT_DIR /
                         f"REPORTE_AUDITORIA_04C_v3_{self.config.TIMESTAMP}.txt")
        with open(reporte_output, 'w', encoding='utf-8') as f:
            f.write('\n'.join(r))

        return r

    # =========================================================================
    # METADATOS ISO 19115
    # =========================================================================

    def generar_metadatos_iso(self):
        """Genera metadatos ISO 19115 para el dataset de exposición."""
        metadatos = {
            "fileIdentifier": f"SCRIPT_04C_v3_EXPOSICION_{self.config.TIMESTAMP}",
            "language": "spa",
            "characterSet": "utf8",
            "hierarchyLevel": "dataset",
            "contact": {
                "individualName": self.config.AUTOR,
                "organisationName": self.config.UNIVERSIDAD,
                "role": "author"
            },
            "dateStamp": self.config.FECHA,
            "identificationInfo": {
                "title": "Exposición agrícola por parroquia — Imbabura",
                "abstract": (
                    "Superficie cosechada (ha) de papa, maíz y fréjol a nivel "
                    "parroquial en la provincia de Imbabura, Ecuador, obtenida "
                    "mediante estadísticas zonales de rasters MapSPAM v2r0 2020 "
                    "sobre límites parroquiales CONALI. Comparación con totales "
                    "provinciales ESPAC 2024. Quinua disponible solo a nivel "
                    "provincial (limitación de fuentes)."
                ),
                "purpose": (
                    "Componente de EXPOSICIÓN del modelo de riesgo agroclimático "
                    "IPCC AR5/AR6."
                ),
                "credit": [
                    "IFPRI/MapSPAM (2024). doi:10.7910/DVN/SWPENT",
                    "CONALI/INEC — Límites parroquiales oficiales Ecuador",
                    "INEC (2024). ESPAC 2024"
                ],
                "spatialResolution": "~10 km (MapSPAM) → parroquial (CONALI)",
                "temporalExtent": "2020 (MapSPAM) / 2024 (ESPAC)",
                "geographicBoundingBox": {
                    "westBoundLongitude": -78.93,
                    "eastBoundLongitude": -77.70,
                    "southBoundLatitude": 0.078,
                    "northBoundLatitude": 0.87,
                },
                "topicCategory": "farming",
                "keyword": [
                    "exposición agrícola", "cultivos andinos", "Imbabura",
                    "desagregación espacial", "MapSPAM", "ESPAC", "CONALI",
                    "papa", "maíz", "fréjol", "quinua"
                ]
            },
            "dataQualityInfo": {
                "scope": "dataset",
                "lineage": {
                    "statement": (
                        "Superficie cosechada extraída de MapSPAM v2r0 2020 "
                        "(IFPRI, doi:10.7910/DVN/SWPENT) mediante estadísticas "
                        "zonales sobre polígonos parroquiales CONALI/INEC. "
                        "Comparación con totales provinciales ESPAC 2024 (INEC). "
                        "Quinua reportada solo a nivel provincial por limitación "
                        "de representatividad muestral de ESPAC (INEC, 2022)."
                    ),
                    "processStep": [
                        {"description": "Descarga automatizada MapSPAM + ESPAC"},
                        {"description": "Carga parroquias CONALI (fuente oficial)"},
                        {"description": "Zonal stats: MapSPAM × parroquias (papa, maíz, fréjol)"},
                        {"description": "Parseo ESPAC: plantada + cosechada por cultivo"},
                        {"description": "Comparación MapSPAM vs ESPAC (sin ajuste forzado)"},
                        {"description": "Quinua documentada como dato provincial únicamente"},
                    ]
                },
                "report": {
                    "n_parroquias": self.resultados.get('n_parroquias', 0),
                    "n_cultivos_parroquiales": 3,
                    "n_cultivos_provinciales": 1,
                    "limitaciones": [
                        "Quinua sin dato parroquial",
                        "MapSPAM ~10km resolución",
                        "Desfase temporal MapSPAM 2020 vs ESPAC 2024"
                    ],
                    "errores": len(self.errores),
                }
            },
            "distributionInfo": {
                "distributionFormat": [
                    {"name": "CSV", "version": "UTF-8-SIG"},
                    {"name": "GeoPackage", "version": "1.3"},
                    {"name": "JSON", "version": "RFC 8259"},
                ]
            },
            "metadataStandardName": "ISO 19115:2014",
            "metadataStandardVersion": "2014",
        }

        metadatos_path = (self.config.METADATOS_DIR /
                         f"ISO19115_SCRIPT_04C_v3_{self.config.TIMESTAMP}.json")
        with open(metadatos_path, 'w', encoding='utf-8') as f:
            json.dump(metadatos, f, indent=2, ensure_ascii=False)
        self._log(f"  ✓ Metadatos ISO: {metadatos_path.name}")

        metadatos_output = (self.config.OUTPUT_DIR /
                           f"metadatos_iso19115_v3_{self.config.TIMESTAMP}.json")
        with open(metadatos_output, 'w', encoding='utf-8') as f:
            json.dump(metadatos, f, indent=2, ensure_ascii=False)

        return metadatos

    # =========================================================================
    # EJECUTAR PIPELINE COMPLETO
    # =========================================================================

    def ejecutar(self):
        """Ejecuta el pipeline completo de exposición agrícola."""
        try:
            # Paso 1: Descarga
            descargas = self.descargar_datos()
            if not descargas.get('parroquias', False):
                raise FileNotFoundError(
                    "Archivo de parroquias CONALI no encontrado. "
                    "Verifique la ruta en ConfiguracionProyecto.PARROQUIAS_FILE"
                )
            if not descargas.get('mapspam', False):
                self._log("⚠ MapSPAM no disponible. Solo se reportará ESPAC.",
                         "WARNING")

            # Paso 2: Cargar parroquias
            self.cargar_parroquias()

            # Paso 3: Zonal stats (solo si MapSPAM disponible)
            if descargas.get('mapspam', False):
                self.calcular_zonal_stats()
            else:
                self.df_mapspam = pd.DataFrame()

            # Paso 4: Parsear ESPAC
            self.parsear_espac()

            # Paso 5: Integrar y validar
            if not self.df_mapspam.empty:
                self.integrar_y_validar()
            else:
                self._log("  Solo datos ESPAC disponibles (nivel provincial).")

            # Reporte y metadatos
            self.generar_reporte_auditoria()
            self.generar_metadatos_iso()

            # Resumen final
            timestamp_fin = datetime.now()
            duracion = (timestamp_fin - self.timestamp_inicio).total_seconds()
            self._log("\n" + "=" * 70)
            self._log("  SCRIPT 04C v3.0 COMPLETADO")
            self._log("=" * 70)
            self._log(f"  Duración total: {duracion:.0f} segundos")
            self._log(f"  Parroquias procesadas: "
                     f"{self.resultados.get('n_parroquias', 0)}")
            self._log(f"  Cultivos parroquiales: papa, maíz, fréjol (MapSPAM)")
            self._log(f"  Cultivos provinciales: quinua (ESPAC)")
            self._log(f"  Errores: {len(self.errores)}")
            self._log(f"  Advertencias: {len(self.advertencias)}")
            self._log(f"  Decisiones documentadas: {len(self.decisiones)}")
            self._log(f"  Productos en: {self.config.OUTPUT_DIR}")
            self._log("=" * 70)

            return True

        except Exception as e:
            self._log(f"\n✗ ERROR CRÍTICO: {e}", "ERROR")
            self._log(traceback.format_exc(), "ERROR")
            return False


# =============================================================================
# EJECUCIÓN PRINCIPAL
# =============================================================================

if __name__ == "__main__":

    print()
    print("=" * 70)
    print("  SCRIPT 04C v3.0 — EXPOSICIÓN AGRÍCOLA PARROQUIAL")
    print("  Riesgo Agroclimático Imbabura")
    print("  Método: exactextract (fracción de cobertura exacta)")
    print("=" * 70)
    print()

    config = ConfiguracionProyecto()

    if not config.BASE_DIR.exists():
        print(f"  ⚠ Ruta base no encontrada: {config.BASE_DIR}")
        print(f"    Ajuste BASE_DIR en ConfiguracionProyecto")
        sys.exit(1)

    pipeline = ExposicionAgricola(config)
    exito = pipeline.ejecutar()

    if exito:
        print(f"\n  ✓ Script 04C v3.0 ejecutado exitosamente.")
        print(f"  Resultados en: {config.OUTPUT_DIR}")
    else:
        print(f"\n  ✗ Script 04C v3.0 completó con errores.")

    print()